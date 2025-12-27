"""
PTSScan Reporter Module
Handles output generation in JSON and Excel formats
"""

import json
import re
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class Reporter:
    """Handles report generation in various formats"""
    
    def __init__(self, results: List[Dict[str, Any]]):
        """
        Initialize reporter with scan results
        
        Args:
            results: List of scan result dictionaries
        """
        self.results = results
        self.timestamp = datetime.now().isoformat()
        self._enhance_results()  # Add line numbers and code snippets
    
    def _extract_line_number(self, violation: str) -> Optional[int]:
        """
        Extract line number from violation message
        Supports various formats: "line 123", "at line 45", "[Line 67]", etc.
        """
        patterns = [
            r'line\s+(\d+)',  # "line 123" or "at line 123"
            r'\[Line\s+(\d+)\]',  # "[Line 123]"
            r'at\s+line\s+(\d+)',  # "at line 123"
            r'on\s+line\s+(\d+)',  # "on line 123"
        ]
        
        for pattern in patterns:
            match = re.search(pattern, violation, re.IGNORECASE)
            if match:
                return int(match.group(1))
        
        return None
    
    def _get_code_snippet(self, filepath: str, line_number: int, context_lines: int = 2) -> Optional[str]:
        """
        Extract code snippet from file around the given line number
        
        Args:
            filepath: Path to the source file
            line_number: Line number to extract
            context_lines: Number of lines before and after to include
        
        Returns:
            Code snippet with line numbers, or None if file cannot be read
        """
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            
            if line_number < 1 or line_number > len(lines):
                return None
            
            # Calculate range (1-indexed)
            start = max(1, line_number - context_lines)
            end = min(len(lines), line_number + context_lines)
            
            # Build snippet with line numbers
            snippet_lines = []
            for i in range(start - 1, end):
                line_num = i + 1
                prefix = ">>>" if line_num == line_number else "   "
                snippet_lines.append(f"{prefix} {line_num:4d} | {lines[i].rstrip()}")
            
            return "\n".join(snippet_lines)
        
        except Exception:
            return None
    
    def _enhance_results(self):
        """
        Enhance results by adding line numbers and code snippets to each bug
        """
        for file_result in self.results:
            filepath = file_result['file']
            
            for bug in file_result['bugs']:
                # Extract line number from violation message
                line_num = self._extract_line_number(bug['violation'])
                bug['line_number'] = line_num
                
                # Get code snippet if line number found
                if line_num:
                    snippet = self._get_code_snippet(filepath, line_num)
                    bug['code_snippet'] = snippet
                else:
                    bug['code_snippet'] = None
    
    def generate_json(self, output_file: str):
        """
        Generate JSON report
        
        Args:
            output_file: Path to output JSON file
        """
        report = {
            'scan_info': {
                'timestamp': self.timestamp,
                'total_files': len(self.results),
                'total_bugs': sum(len(r['bugs']) for r in self.results)
            },
            'results': []
        }
        
        # Process each file's results
        for file_result in self.results:
            file_report = {
                'file': file_result['file'],
                'bugs_count': len(file_result['bugs']),
                'rules_applied': file_result.get('rules_applied', 0),
                'bugs': []
            }
            
            # Add bug details
            for bug in file_result['bugs']:
                bug_entry = {
                    'rule': bug['rule'],
                    'category': bug['category'],
                    'severity': bug['severity'],
                    'description': bug['description'],
                    'violation': bug['violation'],
                    'line_number': bug.get('line_number'),
                    'code_snippet': bug.get('code_snippet')
                }
                file_report['bugs'].append(bug_entry)
            
            report['results'].append(file_report)
        
        # Add summary statistics
        report['summary'] = self._generate_summary()
        
        # Write to file
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(report, f, indent=2, ensure_ascii=False)
    
    def generate_excel(self, output_file: str):
        """
        Generate Excel report with multiple sheets
        
        Args:
            output_file: Path to output Excel file
        """
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create sheets
        self._create_summary_sheet(wb)
        self._create_details_sheet(wb)
        self._create_by_file_sheet(wb)
        self._create_by_category_sheet(wb)
        
        # Save workbook
        wb.save(output_file)
    
    def _create_summary_sheet(self, wb: openpyxl.Workbook):
        """Create summary sheet with overall statistics"""
        ws = wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "PTSScan Security Analysis Report"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells('A1:D1')
        
        # Scan info
        ws['A3'] = "Scan Information"
        ws['A3'].font = Font(size=12, bold=True)
        
        ws['A4'] = "Timestamp:"
        ws['B4'] = self.timestamp
        
        ws['A5'] = "Total Files Scanned:"
        ws['B5'] = len(self.results)
        
        total_bugs = sum(len(r['bugs']) for r in self.results)
        ws['A6'] = "Total Issues Found:"
        ws['B6'] = total_bugs
        
        # Severity breakdown
        ws['A8'] = "Issues by Severity"
        ws['A8'].font = Font(size=12, bold=True)
        
        severity_counts = self._count_by_severity()
        row = 9
        for severity, count in sorted(severity_counts.items()):
            ws[f'A{row}'] = f"{severity.capitalize()}:"
            ws[f'B{row}'] = count
            
            # Color code by severity
            if severity == 'critical':
                ws[f'B{row}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                ws[f'B{row}'].font = Font(color="FFFFFF", bold=True)
            elif severity == 'high':
                ws[f'B{row}'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            elif severity == 'medium':
                ws[f'B{row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            elif severity == 'low':
                ws[f'B{row}'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            
            row += 1
        
        # Category breakdown
        ws[f'A{row+1}'] = "Issues by Category"
        ws[f'A{row+1}'].font = Font(size=12, bold=True)
        
        category_counts = self._count_by_category()
        row += 2
        for category, count in sorted(category_counts.items(), key=lambda x: x[1], reverse=True):
            ws[f'A{row}'] = f"{category.replace('_', ' ').title()}:"
            ws[f'B{row}'] = count
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _create_details_sheet(self, wb: openpyxl.Workbook):
        """Create detailed findings sheet"""
        ws = wb.create_sheet("All Issues")
        
        # Headers
        headers = ['#', 'File', 'Line', 'Category', 'Rule', 'Severity', 'Description', 'Violation Details', 'Code Snippet']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data
        row = 2
        bug_num = 1
        
        for file_result in self.results:
            for bug in file_result['bugs']:
                ws.cell(row=row, column=1, value=bug_num)
                ws.cell(row=row, column=2, value=Path(bug['file']).name)
                
                # Line number
                line_cell = ws.cell(row=row, column=3, value=bug.get('line_number') or 'N/A')
                if bug.get('line_number'):
                    line_cell.font = Font(bold=True, color="0000FF")
                
                ws.cell(row=row, column=4, value=bug['category'].replace('_', ' ').title())
                ws.cell(row=row, column=5, value=bug['rule'])
                
                severity_cell = ws.cell(row=row, column=6, value=bug['severity'].upper())
                # Color code severity
                if bug['severity'] == 'critical':
                    severity_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    severity_cell.font = Font(color="FFFFFF", bold=True)
                elif bug['severity'] == 'high':
                    severity_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                elif bug['severity'] == 'medium':
                    severity_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                elif bug['severity'] == 'low':
                    severity_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                
                ws.cell(row=row, column=7, value=bug['description'])
                ws.cell(row=row, column=8, value=bug['violation'])
                
                # Code snippet with monospace font
                snippet_cell = ws.cell(row=row, column=9, value=bug.get('code_snippet') or 'N/A')
                snippet_cell.font = Font(name='Courier New', size=9)
                snippet_cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                row += 1
                bug_num += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 8   # Line number
        ws.column_dimensions['D'].width = 20  # Category
        ws.column_dimensions['E'].width = 30  # Rule
        ws.column_dimensions['F'].width = 12  # Severity
        ws.column_dimensions['G'].width = 40  # Description
        ws.column_dimensions['H'].width = 50  # Violation
        ws.column_dimensions['I'].width = 60  # Code snippet
        
        # Add filters
        ws.auto_filter.ref = ws.dimensions
    
    def _create_by_file_sheet(self, wb: openpyxl.Workbook):
        """Create sheet showing issues grouped by file"""
        ws = wb.create_sheet("By File")
        
        # Headers
        headers = ['File', 'Total Issues', 'Critical', 'High', 'Medium', 'Low', 'Categories']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data
        row = 2
        for file_result in self.results:
            ws.cell(row=row, column=1, value=Path(file_result['file']).name)
            ws.cell(row=row, column=2, value=len(file_result['bugs']))
            
            # Count by severity
            severity_counts = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0}
            categories = set()
            
            for bug in file_result['bugs']:
                severity_counts[bug['severity']] += 1
                categories.add(bug['category'])
            
            ws.cell(row=row, column=3, value=severity_counts['critical'])
            ws.cell(row=row, column=4, value=severity_counts['high'])
            ws.cell(row=row, column=5, value=severity_counts['medium'])
            ws.cell(row=row, column=6, value=severity_counts['low'])
            ws.cell(row=row, column=7, value=', '.join(sorted(categories)))
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 40
        
        # Add filters
        ws.auto_filter.ref = ws.dimensions
    
    def _create_by_category_sheet(self, wb: openpyxl.Workbook):
        """Create sheet showing issues grouped by category"""
        ws = wb.create_sheet("By Category")
        
        # Headers
        headers = ['Category', 'Total Issues', 'Critical', 'High', 'Medium', 'Low', 'Affected Files']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Aggregate by category
        category_data = {}
        
        for file_result in self.results:
            for bug in file_result['bugs']:
                cat = bug['category']
                if cat not in category_data:
                    category_data[cat] = {
                        'total': 0,
                        'critical': 0,
                        'high': 0,
                        'medium': 0,
                        'low': 0,
                        'files': set()
                    }
                
                category_data[cat]['total'] += 1
                category_data[cat][bug['severity']] += 1
                category_data[cat]['files'].add(Path(bug['file']).name)
        
        # Data
        row = 2
        for category, data in sorted(category_data.items(), key=lambda x: x[1]['total'], reverse=True):
            ws.cell(row=row, column=1, value=category.replace('_', ' ').title())
            ws.cell(row=row, column=2, value=data['total'])
            ws.cell(row=row, column=3, value=data['critical'])
            ws.cell(row=row, column=4, value=data['high'])
            ws.cell(row=row, column=5, value=data['medium'])
            ws.cell(row=row, column=6, value=data['low'])
            ws.cell(row=row, column=7, value=', '.join(sorted(data['files'])))
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 50
        
        # Add filters
        ws.auto_filter.ref = ws.dimensions
    
    def _count_by_severity(self) -> Dict[str, int]:
        """Count issues by severity level"""
        counts = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0}
        
        for file_result in self.results:
            for bug in file_result['bugs']:
                counts[bug['severity']] += 1
        
        return counts
    
    def _count_by_category(self) -> Dict[str, int]:
        """Count issues by category"""
        counts = {}
        
        for file_result in self.results:
            for bug in file_result['bugs']:
                cat = bug['category']
                counts[cat] = counts.get(cat, 0) + 1
        
        return counts
    
    def _generate_summary(self) -> Dict[str, Any]:
        """Generate summary statistics"""
        return {
            'total_files': len(self.results),
            'total_bugs': sum(len(r['bugs']) for r in self.results),
            'by_severity': self._count_by_severity(),
            'by_category': self._count_by_category(),
            'files_with_issues': len([r for r in self.results if r['bugs']]),
            'clean_files': len([r for r in self.results if not r['bugs']])
        }
